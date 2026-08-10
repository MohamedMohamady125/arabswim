"""Repair Hy-Tek 15-char truncated names in Egyptian meet Excel files.

Usage:
    python manage.py repair_egypt_names /path/to/meet_folder [--out DIR]

Reads every PDF in the folder (recursively) to build a full-name index,
then rewrites the Name column of every .xlsx in the folder, saving
repaired copies to <folder>/repaired/ plus a repair_report.txt with
per-file stats and the list of ambiguous / unmatched names.
"""
import glob
import os
from collections import Counter

from django.core.management.base import BaseCommand, CommandError

from importer.egypt_names import (
    NameRepairIndex, extract_pdf_entries, is_truncated_length,
)


class Command(BaseCommand):
    help = 'Expand truncated athlete names in Egyptian meet Excel exports using the companion PDFs'

    def add_arguments(self, parser):
        parser.add_argument('folder', help='Meet folder containing .xlsx files and PDFs')
        parser.add_argument('--out', default='', help='Output directory (default <folder>/repaired)')

    def handle(self, *args, **opts):
        folder = os.path.abspath(opts['folder'])
        if not os.path.isdir(folder):
            raise CommandError(f'Not a folder: {folder}')
        out_dir = opts['out'] or os.path.join(folder, 'repaired')
        os.makedirs(out_dir, exist_ok=True)

        pdfs = sorted(glob.glob(os.path.join(folder, '**', '*.pdf'), recursive=True))
        xlsxs = sorted(f for f in glob.glob(os.path.join(folder, '*.xlsx'))
                       if not f.startswith(out_dir))
        if not xlsxs:
            raise CommandError('No .xlsx files in folder')

        entries = set()
        for p in pdfs:
            found = extract_pdf_entries(p)
            entries |= found
            self.stdout.write(f'  PDF {os.path.basename(p)}: {len(found)} entries')
        index = NameRepairIndex(entries)
        self.stdout.write(self.style.NOTICE(
            f'{len(pdfs)} PDFs → {len(index)} distinct full names'))
        if not len(index):
            raise CommandError('No names extracted from PDFs — nothing to repair with')

        report = [f'Egypt name repair — {os.path.basename(folder)}',
                  f'PDFs: {len(pdfs)}   full names: {len(index)}', '']
        totals = Counter()

        for xf in xlsxs:
            stats, problems = self._repair_workbook(xf, index, out_dir)
            totals.update(stats)
            report.append(f'{os.path.basename(xf)}: '
                          f'{stats["truncated"]} truncated → '
                          f'{stats["repaired"]} repaired, '
                          f'{stats["ambiguous"]} ambiguous, '
                          f'{stats["no_match"]} no match')
            for kind, name, age, team in sorted(problems):
                report.append(f'    [{kind}] {name}  (age {age}, {team})')
            self.stdout.write(report[-1 - len(problems)])

        pct = (totals['repaired'] / totals['truncated'] * 100) if totals['truncated'] else 0
        summary = (f'TOTAL: {totals["truncated"]} truncated names → '
                   f'{totals["repaired"]} repaired ({pct:.0f}%), '
                   f'{totals["ambiguous"]} ambiguous, {totals["no_match"]} no match')
        report += ['', summary]
        report_path = os.path.join(out_dir, 'repair_report.txt')
        with open(report_path, 'w') as fh:
            fh.write('\n'.join(report) + '\n')
        self.stdout.write(self.style.SUCCESS(summary))
        self.stdout.write(f'Repaired files + report in {out_dir}')

    def _repair_workbook(self, path, index, out_dir):
        import openpyxl
        wb = openpyxl.load_workbook(path)
        stats = Counter()
        problems = set()  # (kind, name, age, team)
        for ws in wb.worksheets:
            header = [c.value for c in ws[1]]
            if 'Name' not in header:
                continue
            ncol = header.index('Name') + 1
            acol = header.index('Age') + 1 if 'Age' in header else None
            tcol = header.index('Team') + 1 if 'Team' in header else None
            # Cache per (name, age, team): thousands of rows repeat names
            cache = {}
            for row in ws.iter_rows(min_row=2):
                raw = row[ncol - 1].value
                if not raw or not isinstance(raw, str):
                    continue
                name = raw.strip()
                if not is_truncated_length(name):
                    continue
                age = row[acol - 1].value if acol else None
                age = age if isinstance(age, int) else None
                team = str(row[tcol - 1].value or '') if tcol else ''
                stats['truncated'] += 1
                key = (name, age, team)
                if key not in cache:
                    cache[key] = index.repair(name, age=age, team=team)
                full, status = cache[key]
                if full:
                    row[ncol - 1].value = full
                    stats['repaired'] += 1
                else:
                    stats[status] += 1
                    problems.add((status, name, age, team))
        out_path = os.path.join(out_dir, os.path.basename(path))
        wb.save(out_path)
        return stats, problems
